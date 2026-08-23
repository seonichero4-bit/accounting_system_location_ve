import io
from datetime import date
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, numbers
from django.http import HttpResponse

# Asumiendo importaciones relativas de los modelos
# from data_access.models.purchase_book import PurchaseLedgerInvoice
# from data_access.models.base import FiscalProfile

class PurchaseLedgerExcelBuilder:
    """
    Servicio de un solo propósito para la construcción en memoria del Libro de Compras.
    Aplica separación de operaciones corrientes, ajustes a períodos anteriores y 
    retenciones extemporáneas.
    """
    def __init__(self, fiscal_profile, fiscal_period: date, queryset):
        self.fiscal_profile = fiscal_profile
        self.fiscal_period = fiscal_period
        self.queryset = queryset
        
        # Evaluación del tipo de contribuyente (SPECIAL)
        self.is_special_taxpayer = (self.fiscal_profile.taxpayer_type == 'SPECIAL')
        
        # Configuración del buffer y libro
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Libro de Compras"
        self.buffer = io.BytesIO()
        
        # Estilos recurrentes
        self.bold_font = Font(name='Arial', size=10, bold=True)
        self.normal_font = Font(name='Arial', size=10)
        self.double_bottom_border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='double', color='000000')
        )
        self.money_format = '#,##0.00;(#,##0.00);"-"'
        
        # Diccionario para acumular los totales del resumen
        self.summary = {
            'no_gravadas': Decimal('0.00'),
            'import_gen_base': Decimal('0.00'), 'import_gen_vat': Decimal('0.00'),
            'import_adi_base': Decimal('0.00'), 'import_adi_vat': Decimal('0.00'),
            'import_red_base': Decimal('0.00'), 'import_red_vat': Decimal('0.00'),
            'internal_gen_base': Decimal('0.00'), 'internal_gen_vat': Decimal('0.00'),
            'internal_adi_base': Decimal('0.00'), 'internal_adi_vat': Decimal('0.00'),
            'internal_red_base': Decimal('0.00'), 'internal_red_vat': Decimal('0.00'),
            'ajustes_base': Decimal('0.00'), 'ajustes_vat': Decimal('0.00'),
            'retenciones_periodo': Decimal('0.00'),
            'retenciones_extemporaneas': Decimal('0.00'),
            'Credito fiscal totalmente deducible': Decimal('0.00'),
            'Credito fiscal parcialmente deducible': Decimal('0.00'),
        }

    def _configure_page_setup(self):
        """Ajusta el documento para renderización horizontal en una sola página a lo ancho[cite: 6]."""
        self.sheet.page_setup.orientation = self.sheet.ORIENTATION_LANDSCAPE
        self.sheet.page_setup.paperSize = self.sheet.PAPERSIZE_LETTER
        self.sheet.sheet_properties.pageSetUpPr.fitToPage = True
        self.sheet.page_setup.fitToWidth = 1
        self.sheet.page_setup.fitToHeight = 0

    def _write_header(self):
        """Escribe el membrete fiscal del documento[cite: 6]."""
        self.sheet.append([f"Contribuyente: {self.fiscal_profile.name}", f"RIF: {self.fiscal_profile.rif}"])
        self.sheet.append(["LIBRO DE COMPRAS FISCAL"])
        self.sheet.append([f"Período Fiscal: {self.fiscal_period.strftime('%m-%Y')}"])
        self.sheet.append([]) # Línea vacía

    def _write_table_headers(self) -> int:
        """Construye dinámicamente los encabezados. Retorna la longitud de columnas[cite: 6]."""
        headers = [
            "N° Operación", "Fecha Documento", "N° R.I.F.", 
            "Nombre o Razón Social", "Número de Documento", "Número de Control",
            "N° Control Nota Débito", "N° Control Nota Crédito"
        ]
        
        if self.is_special_taxpayer:
            headers.append("N° Comprobante Retención IVA")
            headers.append("Fecha de Aplicacion Retención IVA")

        headers.extend([
            "Tipo de Transacción", "N° Documento Afectado", "N° Planilla Importación",
            "N° Expediente Importación", "Total Compras Con IVA", "Compras Exentas",
            "Compras Exoneradas", "Compras No Sujetas", "Sin Derecho a Crédito",
            "Base Imponible", "% Alícuota", "Impuesto IVA"
        ])
        
        if self.is_special_taxpayer:
            headers.append("Total IVA Retenido")

        self.sheet.append(headers)
        
        # Aplicar negrita a los encabezados
        for cell in self.sheet[self.sheet.max_row]:
            cell.font = self.bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        return len(headers)

    def _process_invoice_row(self, invoice, index: int, is_adjustment: bool = False):
        """Mapea una instancia de PurchaseLedgerInvoice a una fila de Excel."""
        row = [
            index,
            invoice.date.strftime('%d/%m/%Y') if invoice.date else "",
            invoice.supplier.rif,
            invoice.supplier.name,
            invoice.number,
            invoice.invoice_control if invoice.document_type == 'INVOICE' else "",
            invoice.invoice_control if invoice.document_type == 'DEBIT_NOTE' else "",
            invoice.invoice_control if invoice.document_type == 'CREDIT_NOTE' else "",
        ]

        if self.is_special_taxpayer:
            # Obtener número de comprobante si existe y está procesado
            cert = getattr(invoice, 'vat_withholding_certificate', None)
            row.append(cert.document_number if cert else "")
            row.append(cert.application_date if cert else "")

        row.extend([
            invoice.get_transaction_type_display(),
            invoice.affected_invoice.number if invoice.affected_invoice else "",
            invoice.import_form_number or "",
            invoice.import_file_number or "",
            invoice.total_purchase,
            invoice.exempt_amount,
            invoice.amount_exonerated,
            invoice.amount_not_subject,
            invoice.amount_without_right_to_credit,
            invoice.taxable_base,
            invoice.get_vat_percentage_display(),
            invoice.vat_amount
        ])

        if self.is_special_taxpayer:
            cert = getattr(invoice, 'vat_withholding_certificate', None)
            row.append(cert.vat_withheld_amount if cert else Decimal('0.00'))

        # Actualizar totales para el resumen
        self._update_summary_totals(invoice, is_adjustment)
        
        return row

    def _process_late_withholding_row(self, cert, index: int):
        """
        Mapea una instancia de VatWithholdingCertificate extemporáneo a una fila de Excel.
        Acumula únicamente el monto de la retención para el período actual.
        """
        invoice = cert.purchase_invoice
        
        row = [
            index,
            invoice.date.strftime('%d/%m/%Y') if invoice.date else "",
            invoice.supplier.rif,
            invoice.supplier.name,
            invoice.number,
            invoice.invoice_control,
            "",
            "",

            cert.document_number,
            cert.application_date.strftime('%d/%m/%Y'),

            invoice.get_transaction_type_display(),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            cert.vat_withheld_amount,
        ]

        if self.is_special_taxpayer:
            # Acumula al resumen de forma independiente, ya que la base/IVA de la factura 
            # no afecta los créditos fiscales del período actual (fueron declarados antes)
            sign = Decimal('-1') if invoice.document_type == 'CREDIT_NOTE' else Decimal('1')
            withheld = cert.vat_withheld_amount or Decimal('0.00')
            row.append(withheld)
            self.summary['retenciones_extemporaneas'] += (withheld * sign)

        return row

    def _update_summary_totals(self, invoice, is_adjustment: bool):
        """Alimenta los agregadores del resumen financiero en memoria."""
        # Factor de signo dinámico para cualquier operación (corriente o ajuste)
        sign = Decimal('-1') if invoice.document_type == 'CREDIT_NOTE' else Decimal('1')

        # Acumulación de Deducibilidad de Crédito Fiscal
        if invoice.deductibility == "Deducible":
            self.summary['Credito fiscal totalmente deducible'] += (invoice.vat_amount * sign)
        elif invoice.deductibility == "Parcialmente deducible":
            self.summary['Credito fiscal parcialmente deducible'] += (invoice.vat_amount * sign)

        if is_adjustment:
            self.summary['ajustes_base'] += (invoice.taxable_base * sign)
            self.summary['ajustes_vat'] += (invoice.vat_amount * sign)
            return

        # Acumular conceptos no gravados con su signo correspondiente
        self.summary['no_gravadas'] += (
            invoice.exempt_amount + invoice.amount_exonerated + 
            invoice.amount_not_subject + invoice.amount_without_right_to_credit
        ) * sign

        prefix = "import" if invoice.purchase_type == 'IMPORT' else "internal"
        
        if invoice.vat_percentage == 1:
            suffix = "gen"
        elif invoice.vat_percentage == 2:
            suffix = "red"
        elif invoice.vat_percentage == 3:
            suffix = "adi"
        else:
            return

        # Aplica el signo a las operaciones corrientes del período
        self.summary[f'{prefix}_{suffix}_base'] += (invoice.taxable_base * sign)
        self.summary[f'{prefix}_{suffix}_vat'] += (invoice.vat_amount * sign)

       # Acumular IVA retenido en operaciones corrientes del periodo fiscal
        if self.is_special_taxpayer:
            cert = getattr(invoice, 'vat_withholding_certificate', None)
            if cert and cert.vat_withheld_amount:
                self.summary['retenciones_periodo'] += (cert.vat_withheld_amount * sign)

    def _write_operations_data(self, total_cols: int):
        """Segmenta y renderiza las operaciones del período, ajustes y extemporáneas[cite: 6]."""
        current_ops = []
        adjustments = []
        late_vat_withholding = []
        
        # Clasificación inicial en memoria
        for invoice in self.queryset:
            if invoice.vat_withholding_certificate and invoice.vat_withholding_certificate.fiscal_period < self.fiscal_period:
                late_vat_withholding.append(invoice.vat_withholding_certificate)
            if invoice.affected_invoice and invoice.affected_invoice.fiscal_period < self.fiscal_period:
                adjustments.append(invoice)
            else:
                current_ops.append(invoice)

        # 1. Operaciones Corrientes
        index = 1
        for inv in current_ops:
            self.sheet.append(self._process_invoice_row(inv, index, is_adjustment=False))
            index += 1

        # Subtotal de operaciones corrientes (Aplica estilos)
        self.sheet.append([""] * total_cols)
        
        # 2. Ajustes a Períodos Anteriores
        if adjustments:
            self.sheet.append(["AJUSTES A CRÉDITOS FISCALES DE PERÍODOS ANTERIORES"])
            for inv in adjustments:
                self.sheet.append(self._process_invoice_row(inv, index, is_adjustment=True))
                index += 1

        # 3. Retenciones Extemporáneas
        if late_vat_withholding:
            # Añadir separación con la tabla anterior si existen registros
            self.sheet.append([""] * total_cols)
            
            # Título de la sección
            self.sheet.append(["RETENCIONES DE IVA EXTEMPORÁNEAS (PERÍODOS ANTERIORES)"])
            
            # Mapeo e inserción de filas
            for cert in late_vat_withholding:
                self.sheet.append(self._process_late_withholding_row(cert, index))
                index += 1

    def _write_summary_section(self):
        """Construye el bloque consolidado del Resumen de Créditos Fiscales."""
        self.sheet.append([])
        self.sheet.append([])
        self.sheet.append(["RESUMEN DEL LIBRO DE COMPRAS"])
        
        headers = ["Concepto", "Base Imponible", "Crédito Fiscal / IVA Retenido"]
        self.sheet.append(headers)

        rows = [
            ("Compras no Gravadas y/o sin Derecho a Crédito Fiscal", self.summary['no_gravadas'], "N/A"),
            ("Importaciones Gravadas por Alícuota General", self.summary['import_gen_base'], self.summary['import_gen_vat']),
            ("Importaciones Gravadas por Alícuota General + Adicional", self.summary['import_adi_base'], self.summary['import_adi_vat']),
            ("Importaciones Gravadas por Alícuota Reducida", self.summary['import_red_base'], self.summary['import_red_vat']),
            ("Compras Internas Gravadas por Alícuota General", self.summary['internal_gen_base'], self.summary['internal_gen_vat']),
            ("Compras Internas Gravadas por Alícuota General + Adicional", self.summary['internal_adi_base'], self.summary['internal_adi_vat']),
            ("Compras Internas Gravadas por Alícuota Reducida", self.summary['internal_red_base'], self.summary['internal_red_vat']),         
            ("Total Compras y Créditos Fiscales del Período", "N/A", 
             self.summary['no_gravadas'] + 
             self.summary['import_gen_base'] +
             self.summary['import_adi_base'] +
             self.summary['import_red_base'] +
             self.summary['internal_gen_base'] +
             self.summary['internal_adi_base'] +
             self.summary['internal_red_base'],

             self.summary['import_gen_vat'] +
             self.summary['import_adi_vat'] +
             self.summary['import_red_vat'] +
             self.summary['internal_gen_vat'] +
             self.summary['internal_adi_vat'] +
             self.summary['internal_red_vat']
            ),
           
            ("Crédito fiscal totalmente deducible", "N/A", self.summary['Credito fiscal totalmente deducible']),
            ("Crédito fiscal parcialmente deducible", "N/A", self.summary['Credito fiscal parcialmente deducible']),
        ]

        # Agregar los acumulados de retenciones solo si es contribuyente especial
        if self.is_special_taxpayer:
            self.sheet.append([])
            rows.extend([
                ("Total IVA Retenido en Operaciones del Período", "N/A", self.summary['retenciones_periodo']),
                ("Total IVA Retenido en Operaciones Extemporáneas", "N/A", self.summary['retenciones_extemporaneas'])
            ])

        self.sheet.append([])
        ("Ajuste a los créditos fiscales de períodos anteriores", self.summary['ajustes_base'], self.summary['ajustes_vat']),

        for row in rows:
            self.sheet.append(row)

    def _apply_styles_and_formatting(self):
        """Asigna formatos monetarios y el doble subrayado contable exigido[cite: 6]."""
        for row in self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row):
            for cell in row:
                if isinstance(cell.value, Decimal) or isinstance(cell.value, float):
                    cell.number_format = self.money_format

    def build(self) -> 'PurchaseLedgerExcelBuilder':
        """Orquesta la construcción del documento secuencialmente[cite: 6]."""
        self._configure_page_setup()
        self._write_header()
        total_cols = self._write_table_headers()
        self._write_operations_data(total_cols)
        self._write_summary_section()
        self._apply_styles_and_formatting()
        return self

    def get_response(self) -> HttpResponse:
        """Persiste el archivo en el BytesIO buffer y emite el payload HTTP[cite: 6]."""
        self.workbook.save(self.buffer)
        self.buffer.seek(0)
        
        filename = f"Libro_Compras_{self.fiscal_profile.rif}_{self.fiscal_period.strftime('%Y%m')}.xlsx"
        
        response = HttpResponse(
            self.buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response