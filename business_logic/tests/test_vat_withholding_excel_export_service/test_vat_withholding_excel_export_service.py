"""Módulo de pruebas unitarias para el servicio de exportación de retenciones de IVA en Excel."""

import io
from datetime import date
from decimal import Decimal
from typing import Any

import openpyxl
import pytest

from data_access.models.base import FiscalProfile
from data_access.models.purchase_book import PurchaseLedgerInvoice
from data_access.models.supplier import LocalSupplier
from data_access.models.vat_withholding import VatWithholdingCertificate
from business_logic.services.vat_withholding_excel_export_service import VatWithholdingExcelExportService


def test_id_hp_001_generate_excel_stream_with_invoice(
    db: Any, 
    fiscal_profile: FiscalProfile, 
    vat_withholding_certificate: VatWithholdingCertificate
) -> None:
    """Valida la generación del excel con un comprobante asociado a una factura regular (INVOICE)."""
    # Arrange
    service = VatWithholdingExcelExportService(
        fiscal_profile=fiscal_profile,
        fiscal_period=date(2026, 8, 15)
    )

    # Act
    excel_stream = service.generate_excel_stream()

    # Assert
    assert isinstance(excel_stream, io.BytesIO)
    assert excel_stream.tell() == 0

    wb = openpyxl.load_workbook(filename=excel_stream)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    assert len(rows) == 1
    row = rows[0]
    
    assert row[0] == "J123456780"           # Col A: RIF Perfil
    assert row[1] == "202608"                 # Col B: Periodo
    assert str(row[2])[:10] == "2026-08-10"   # Col C: Fecha de Factura ajustada
    assert row[3] == "C"                      # Col D: Constante C
    assert row[4] == "01"                     # Col E: 01 = INVOICE
    assert row[5] == "J123456780"             # Col F: RIF Proveedor
    assert row[11] == "0"                     # Col L: affected_invoice nulo
    assert float(row[14]) == 16.00            # Col O: Alícuota
    assert row[15] == "0"                     # Col P: import_file_number nulo


def test_id_hp_002_generate_excel_stream_with_debit_note(
    db: Any, 
    fiscal_profile: FiscalProfile, 
    local_supplier: LocalSupplier,
    purchase_ledger_invoice: PurchaseLedgerInvoice
) -> None:
    """Confirma el mapeo de tipos de documento de nota de débito con factura afectada y expediente."""
    # Arrange
    purchase_ledger_invoice.number = "FACT-0099"
    purchase_ledger_invoice.save()

    debit_note = PurchaseLedgerInvoice(
        fiscal_profile=fiscal_profile,
        transaction_type=PurchaseLedgerInvoice.TransactionType.REGISTRO,
        document_type=PurchaseLedgerInvoice.DocumentType.DEBIT_NOTE,
        number="00004568",
        invoice_control="00-00124",
        supplier=local_supplier,
        date=date(2026, 8, 11),
        fiscal_period=date(2026, 8, 15),
        purchase_type=PurchaseLedgerInvoice.PurchaseType.IMPORT,
        import_form_number="FORM-2026-01",
        import_file_number="EXP-2026-01",
        affected_invoice=purchase_ledger_invoice,
        taxable_base=Decimal("1000.00"),
        vat_percentage=PurchaseLedgerInvoice.VatPercentageChoices.GENERAL,
        vat_amount=Decimal("160.00"),
        total_purchase=Decimal("1160.00"),
        status=PurchaseLedgerInvoice.InvoiceStatus.PRELIMINARY,
        invoicecategory=PurchaseLedgerInvoice.InvoiceCategory.INVENTARIO,
    )
    debit_note.full_clean()
    debit_note.save()

    cert = VatWithholdingCertificate(
        fiscal_profile=fiscal_profile,
        purchase_invoice=debit_note,
        application_date=date(2026, 8, 11),
        fiscal_period=date(2026, 8, 15),
        vat_withholding_percentage=VatWithholdingCertificate.VatWithholdingChoices.SETENTA_Y_CINCO,
        document_number="202608000002",
        status=VatWithholdingCertificate.CertificateStatus.PRELIMINARY
    )
    cert.full_clean()
    cert.save()

    service = VatWithholdingExcelExportService(
        fiscal_profile=fiscal_profile,
        fiscal_period=date(2026, 8, 15)
    )

    # Act
    excel_stream = service.generate_excel_stream()

    # Assert
    wb = openpyxl.load_workbook(filename=excel_stream)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    
    debit_row = next(row for row in rows if row[4] == "02")
    assert debit_row[4] == "02"            
    assert debit_row[11] == "FACT-0099"    
    assert debit_row[15] == "EXP-2026-01"  


def test_id_hp_003_generate_excel_stream_with_credit_note(
    db: Any, 
    fiscal_profile: FiscalProfile, 
    local_supplier: LocalSupplier,
    purchase_ledger_invoice: PurchaseLedgerInvoice
) -> None:
    """Verifica la correcta transformación de comprobantes asociados a notas de crédito."""
    # Arrange
    purchase_ledger_invoice.number = "FACT-0050"
    purchase_ledger_invoice.save()

    credit_note = PurchaseLedgerInvoice(
        fiscal_profile=fiscal_profile,
        transaction_type=PurchaseLedgerInvoice.TransactionType.REGISTRO,
        document_type=PurchaseLedgerInvoice.DocumentType.CREDIT_NOTE,
        number="00004569",
        invoice_control="00-00125",
        supplier=local_supplier,
        date=date(2026, 8, 12),
        fiscal_period=date(2026, 8, 15),
        purchase_type=PurchaseLedgerInvoice.PurchaseType.INTERNAL,
        affected_invoice=purchase_ledger_invoice,
        taxable_base=Decimal("1000.00"),
        vat_percentage=PurchaseLedgerInvoice.VatPercentageChoices.GENERAL,
        vat_amount=Decimal("160.00"),
        total_purchase=Decimal("1160.00"),
        status=PurchaseLedgerInvoice.InvoiceStatus.PRELIMINARY,
        invoicecategory=PurchaseLedgerInvoice.InvoiceCategory.INVENTARIO,
    )
    credit_note.full_clean()
    credit_note.save()

    cert = VatWithholdingCertificate(
        fiscal_profile=fiscal_profile,
        purchase_invoice=credit_note,
        application_date=date(2026, 8, 12),
        fiscal_period=date(2026, 8, 15),
        vat_withholding_percentage=VatWithholdingCertificate.VatWithholdingChoices.SETENTA_Y_CINCO,
        document_number="202608000003",
        status=VatWithholdingCertificate.CertificateStatus.PRELIMINARY
    )
    cert.full_clean()
    cert.save()

    service = VatWithholdingExcelExportService(
        fiscal_profile=fiscal_profile,
        fiscal_period=date(2026, 8, 15)
    )

    # Act
    excel_stream = service.generate_excel_stream()

    # Assert
    wb = openpyxl.load_workbook(filename=excel_stream)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    
    credit_row = next(row for row in rows if row[4] == "03")
    assert credit_row[4] == "03"         
    assert credit_row[11] == "FACT-0050" 


def test_id_hp_004_sequential_export_of_multiple_records(
    db: Any, 
    fiscal_profile: FiscalProfile, 
    vat_withholding_certificate: VatWithholdingCertificate,
    local_supplier: LocalSupplier
) -> None:
    """Comprueba que el componente mantenga el orden de iteración al procesar múltiples comprobantes."""
    # Arrange
    invoice_2 = PurchaseLedgerInvoice.objects.create(
        fiscal_profile=fiscal_profile,
        transaction_type=PurchaseLedgerInvoice.TransactionType.REGISTRO,
        document_type=PurchaseLedgerInvoice.DocumentType.INVOICE,
        number="00004599",
        invoice_control="00-00199",
        supplier=local_supplier,
        date=date(2026, 8, 14),
        fiscal_period=date(2026, 8, 15),
        purchase_type=PurchaseLedgerInvoice.PurchaseType.INTERNAL,
        taxable_base=Decimal("2000.00"),
        vat_percentage=PurchaseLedgerInvoice.VatPercentageChoices.GENERAL,
        vat_amount=Decimal("320.00"),
        total_purchase=Decimal("2320.00"),
        status=PurchaseLedgerInvoice.InvoiceStatus.PRELIMINARY,
        invoicecategory=PurchaseLedgerInvoice.InvoiceCategory.INVENTARIO,
    )
    VatWithholdingCertificate.objects.create(
        fiscal_profile=fiscal_profile,
        purchase_invoice=invoice_2,
        application_date=date(2026, 8, 14),
        fiscal_period=date(2026, 8, 15),
        vat_withholding_percentage=VatWithholdingCertificate.VatWithholdingChoices.SETENTA_Y_CINCO,
        document_number="202608000004",
        status=VatWithholdingCertificate.CertificateStatus.PRELIMINARY
    )

    service = VatWithholdingExcelExportService(
        fiscal_profile=fiscal_profile,
        fiscal_period=date(2026, 8, 15)
    )

    # Act
    excel_stream = service.generate_excel_stream()

    # Assert
    wb = openpyxl.load_workbook(filename=excel_stream)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    
    assert len(rows) == 2
    assert rows[0][12] == "202608000001" 
    assert rows[1][12] == "202608000004" 


def test_id_ec_001_buffer_pointer_reset_to_zero(
    db: Any, 
    fiscal_profile: FiscalProfile, 
    vat_withholding_certificate: VatWithholdingCertificate
) -> None:
    """Garantiza que el puntero del stream BytesIO sea restablecido al inicio antes de la devolución."""
    # Arrange
    service = VatWithholdingExcelExportService(
        fiscal_profile=fiscal_profile,
        fiscal_period=date(2026, 8, 15)
    )

    # Act
    excel_stream = service.generate_excel_stream()

    # Assert
    assert excel_stream.tell() == 0
    content = excel_stream.read()
    assert len(content) > 0


def test_id_ec_002_decimal_precision_and_zero_handling(
    db: Any, 
    fiscal_profile: FiscalProfile, 
    vat_withholding_certificate: VatWithholdingCertificate
) -> None:
    """Evalúa la precisión en la conversión de montos Decimal hacia las celdas del libro."""
    # Arrange
    service = VatWithholdingExcelExportService(
        fiscal_profile=fiscal_profile,
        fiscal_period=date(2026, 8, 15)
    )

    # Act
    excel_stream = service.generate_excel_stream()

    # Assert
    wb = openpyxl.load_workbook(filename=excel_stream)
    ws = wb.active
    row = list(ws.iter_rows(values_only=True))[0]
    
    assert float(row[8]) == 1160.00
    assert float(row[9]) == 1000.00
    assert float(row[10]) == 120.00
    assert float(row[13]) == 0.00
    assert float(row[14]) == 16.00


def test_id_ec_003_strict_absence_of_visual_styles(
    db: Any, 
    fiscal_profile: FiscalProfile, 
    vat_withholding_certificate: VatWithholdingCertificate
) -> None:
    """Garantiza el cumplimiento estricto de la restricción de diseño que prohíbe estilos personalizados."""
    # Arrange
    service = VatWithholdingExcelExportService(
        fiscal_profile=fiscal_profile,
        fiscal_period=date(2026, 8, 15)
    )

    # Act
    excel_stream = service.generate_excel_stream()

    # Assert
    wb = openpyxl.load_workbook(filename=excel_stream)
    ws = wb.active
    
    for cell in ws[1]:
        assert cell.has_style is False or cell.fill.fill_type is None
        assert cell.font.b is False  


def test_id_ec_004_preservation_of_leading_zeros_and_formats(
    db: Any, 
    fiscal_profile: FiscalProfile, 
    local_supplier: LocalSupplier,
    purchase_ledger_invoice: PurchaseLedgerInvoice,
    vat_withholding_certificate: VatWithholdingCertificate
) -> None:
    """Previene la pérdida de ceros iniciales o formato en identificadores fiscales y documentos."""
    # Arrange
    local_supplier.rif = "J000123456"
    local_supplier.save()
    
    purchase_ledger_invoice.number = "00004567"
    purchase_ledger_invoice.invoice_control = "00-00123"
    purchase_ledger_invoice.save()
    
    service = VatWithholdingExcelExportService(
        fiscal_profile=fiscal_profile,
        fiscal_period=date(2026, 8, 15)
    )

    # Act
    excel_stream = service.generate_excel_stream()

    # Assert
    wb = openpyxl.load_workbook(filename=excel_stream)
    ws = wb.active
    row = list(ws.iter_rows(values_only=True))[0]
    
    assert row[5] == "J000123456" 
    assert row[6] == "00004567"   
    assert row[7] == "00-00123"   


def test_id_ec_005_instantiation_with_null_dependencies_raises_error(
    db: Any
) -> None:
    """Evalúa el comportamiento cuando los atributos del constructor no cumplen con el contrato."""
    # Arrange & Act & Assert
    with pytest.raises((ValueError, TypeError, AttributeError, Exception)):
        service = VatWithholdingExcelExportService(
            fiscal_profile=None,
            fiscal_period=None
        )
        service.generate_excel_stream()